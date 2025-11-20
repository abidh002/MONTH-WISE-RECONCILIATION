import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Invoice Reconciliation Tool", layout="wide")
st.title("💰 Invoice Reconciliation Tool")

st.markdown("### Step 1: Download Templates")

col1, col2 = st.columns(2)

with col1:
    st.download_button(
        "📥 Download Submission Template",
        data="Month,Invoice,Member ID,Transaction Date,Amount\n2025-01,INV001,12345,22-10-2025,1000\n2025-02,INV002,56789,23-10-2025,2000\n2025-03,INV003,11111,24-10-2025,1500",
        file_name="submission_template.csv",
        mime="text/csv"
    )

with col2:
    st.download_button(
        "📥 Download Remittance Template",
        data="Invoice,Payment Reference,Settlement Date,Amount\nINV001,REF001,24-10-2025,800\nINV002,REF002,25-10-2025,2000",
        file_name="remittance_template.csv",
        mime="text/csv"
    )

st.markdown("### Step 2: Upload Your Files")

sub_file = st.file_uploader("Upload Submission File", type=["xlsx", "csv"])
rem_file = st.file_uploader("Upload Remittance File", type=["xlsx", "csv"])

if sub_file and rem_file:
    try:
        # Read submission
        if sub_file.name.lower().endswith(".csv"):
            submission_df = pd.read_csv(sub_file, dtype=str)
        else:
            submission_df = pd.read_excel(sub_file, dtype=str)

        # Read remittance
        if rem_file.name.lower().endswith(".csv"):
            remittance_df = pd.read_csv(rem_file, dtype=str)
        else:
            remittance_df = pd.read_excel(rem_file, dtype=str)

        # Clean column names
        submission_df.columns = submission_df.columns.str.strip().str.title()
        remittance_df.columns = remittance_df.columns.str.strip().str.title()

        # Validate required columns
        required_sub_cols = {"Month", "Invoice", "Member Id", "Transaction Date", "Amount"}
        required_rem_cols = {"Invoice", "Payment Reference", "Settlement Date", "Amount"}

        if not required_sub_cols.issubset(submission_df.columns):
            st.error("❌ Submission file must have columns: Month, Invoice, Member ID, Transaction Date, Amount")
        elif not required_rem_cols.issubset(remittance_df.columns):
            st.error("❌ Remittance file must have columns: Invoice, Payment Reference, Settlement Date, Amount")
        else:
            # Convert Amounts to numeric
            submission_df["Amount"] = pd.to_numeric(submission_df["Amount"], errors="coerce").fillna(0)
            remittance_df["Amount"] = pd.to_numeric(remittance_df["Amount"], errors="coerce").fillna(0)

            # Convert dates
            submission_df["Transaction Date"] = pd.to_datetime(submission_df["Transaction Date"], dayfirst=True, errors='coerce').dt.date
            remittance_df["Settlement Date"] = pd.to_datetime(remittance_df["Settlement Date"], dayfirst=True, errors='coerce').dt.date

            # Aggregate submission by invoice
            submission_agg = submission_df.groupby(["Month", "Invoice"], as_index=False)["Amount"].sum()
            submission_agg.rename(columns={"Amount": "Submission_Amount"}, inplace=True)

            # Aggregate remittance by invoice
            remittance_agg = remittance_df.groupby(["Invoice", "Payment Reference", "Settlement Date"], as_index=False)["Amount"].sum()
            remittance_agg.rename(columns={"Amount": "Remittance_Amount"}, inplace=True)

            # LEFT JOIN: keep all submission records
            result = pd.merge(
                submission_agg,
                remittance_agg,
                on="Invoice",
                how="left"
            )

            # Fill missing remittance info
            result["Remittance_Amount"] = result["Remittance_Amount"].fillna(0)
            result["Payment Reference"] = result["Payment Reference"].fillna("Pending")
            result["Settlement Date"] = result["Settlement Date"].astype(str).replace("NaT", "")

            # Difference
            result["Difference"] = result["Submission_Amount"] - result["Remittance_Amount"]

            # Status
            def get_status(row):
                if row["Remittance_Amount"] == 0:
                    return "Pending from Insurance"
                elif row["Difference"] == 0:
                    return "Matched"
                else:
                    return "Not Matched"

            result["Status"] = result.apply(get_status, axis=1)

            # Reorder columns
            result = result[[
                "Month", "Invoice", "Payment Reference", "Settlement Date",
                "Submission_Amount", "Remittance_Amount", "Difference", "Status"
            ]]

            # Display in Streamlit
            st.markdown("### 🔍 Reconciliation Result (All Submission Data)")
            def highlight_status(row):
                if row["Status"] == "Matched":
                    return ['background-color: #d4edda'] * len(row)  # Light Green
                else:
                    return ['background-color: #fff3cd'] * len(row)  # Light Orange for Not Matched / Pending

            st.dataframe(result.style.apply(highlight_status, axis=1), use_container_width=True)

            # Download Excel with color
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                result.to_excel(writer, index=False, sheet_name="Reconciliation")
            buffer.seek(0)

            # Apply color formatting
            wb = load_workbook(buffer)
            ws = wb.active

            green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
            orange_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")

            for row in range(2, ws.max_row + 1):
                status = ws.cell(row=row, column=8).value
                fill = green_fill if status == "Matched" else orange_fill
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill

            export_buffer = BytesIO()
            wb.save(export_buffer)
            export_buffer.seek(0)

            st.download_button(
                label="📊 Download Reconciliation Result (All Submission Data, Colored)",
                data=export_buffer.getvalue(),
                file_name="Reconciliation_All_Submission.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"❌ Error: {e}")
        st.exception(e)

